# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 09:12:29 2021

@author: Calvin Brown
"""
from openpyxl import Workbook

def readInputs(filename):
    f = open(filename, "r")
    lines = f.readlines()
    constants = lines[0].strip().split(",")
    doses_per_vial = int(constants[0])
    time_between_doses = int(constants[1])
    daily_supplies = list(map(lambda x: int(x), lines[1].strip().split(",")))
    days = len(daily_supplies)
    center_capacities = list(map(lambda x: int(x), lines[2].strip().split(",")))
    individuals = list(map(lambda x: Ind(x[0], float(x[1])), map(lambda x: x.strip().split(","), lines[3:])))
    return AllocationInputs(daily_supplies, days, doses_per_vial, time_between_doses, center_capacities, individuals)

def writeSolution(filename, allocation_dict):
    wb = Workbook()
    days = sorted(allocation_dict.keys())
    for d in days:
        ws = wb.create_sheet("Day_{}\n".format(d))
        centers = sorted(allocation_dict[d].keys())
        for c in centers:
            ws.cell(row=1, column=c+1, value="Center_{}".format(c))
            individuals = allocation_dict[d][c]
            for i in range(len(individuals)):
                ws.cell(row=i+2, column=c+1, value=str(individuals[i]))
    del wb["Sheet"]
    wb.save(filename)
    
def find_objective(allocation_dict):
    objective = 0
    days = sorted(allocation_dict.keys())
    total_days = len(days)
    for d in days:
        centers = sorted(allocation_dict[d].keys())
        for c in centers:
            individuals = allocation_dict[d][c]
            for ind in individuals:
                objective += ind.risk_score * ((2 * total_days - d)/(2 * total_days))
    return objective

class Ind:
    def __init__(self, ind_id, risk_score):
        self.ind_id = ind_id
        self.risk_score = risk_score
    
    def set_score(self, new_score):
        self.risk_score = new_score

    def __lt__(self, other):
        if (self.risk_score > other.risk_score):
            return True
        return False
    
    def __eq__(self, other):
        if (self.risk_score == other.risk_score):
            return True
        return False

    def __repr__(self):
        return "I:" + self.ind_id + "-R:" + str(self.risk_score)

class AllocationInputs:
    def __init__(self, ds, days, dpv, tbd, ccs, inds):
        self.ds = ds
        self.days = days
        self.dpv = dpv
        self.two_dose = True if tbd > 0 else False
        self.tbd = tbd
        self.ccs = ccs
        self.inds = inds

    def __repr__(self):
        dose_repr = "two dose, time between doses: {}".format(self.tbd) if self.two_dose else "one dose"
        ds_repr = self.ds if len(self.ds) <= 10 else str(self.ds[:10]).strip(']') + "..."
        ccs_repr = self.ccs if (len(self.ccs) <= 10) else str(self.ccs[:10]).strip(']') + "..."
        inds_repr = self.inds if len(self.inds) <= 10 else str(self.inds[:10]).strip(']') + "..."
        
        return "Days: {}, doses per vial: {}, {}\nDaily supplies: {}\nCenter capacities: {} \nIndividuals: {}"\
            .format(self.days, self.dpv, dose_repr, ds_repr, ccs_repr, inds_repr)
    
    def write(self, filename):
        f = open(filename, 'w')
        f.write('{},{}\n'.format(self.dpv, self.tbd))
        f.write(','.join(list(map(str, self.ds)))+"\n")
        f.write(','.join(list(map(str, self.ccs)))+"\n")
        f.write('\n'.join(list(map(lambda x: '{},{}'.format(x.ind_id,x.risk_score), self.inds))))
        f.close()
        
if __name__ == "__main__":
    print(readInputs("input/allocation_problem_3.txt"))
    readInputs("input/allocation_problem_3.txt").write("input/test1.txt")
    print(readInputs("input/test1.txt"))
